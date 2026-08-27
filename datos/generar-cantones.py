#!/usr/bin/env python3
"""Genera datos/cantones.csv a partir del archivo de resultados de la
Estimación de Población y Vivienda 2022 del INEC.

Descarga el XLSX del INEC a un directorio temporal y lee dos de sus
cuadros: el 11 (población por provincia, cantón y distrito) y el 12
(extensión territorial en km², misma jerarquía). El archivo del INEC usa
la división territorial de 2022, con 82 cantones: Monteverde y Puerto
Jiménez, creados como cantones en 2021 a partir de los distritos del
mismo nombre de Puntarenas y Golfito, figuran todavía como distritos. El
script los reclasifica como cantones y resta sus cifras a los cantones
de origen, con lo que resultan los 84 cantones actuales.

Uso: python3 datos/generar-cantones.py   (requiere openpyxl)
"""

import csv
import io
import urllib.request
from pathlib import Path

import openpyxl

URL = ("https://admin.inec.cr/sites/default/files/2023-11/"
       "reResultadosEstimacionPoblacionVivienda2022_3.xlsx")
SALIDA = Path(__file__).with_name("cantones.csv")

# Provincias en el orden oficial (el mismo del cuadro del INEC)
PROVINCIAS = ["San José", "Alajuela", "Cartago", "Heredia", "Guanacaste",
              "Puntarenas", "Limón"]

# Distritos que pasaron a ser cantones en 2021: (provincia, cantón de origen, distrito)
NUEVOS_CANTONES = [("Puntarenas", "Puntarenas", "Monteverde"),
                   ("Puntarenas", "Golfito", "Puerto Jiménez")]


def leer_cuadro(libro, hoja, columna):
    """Lee un cuadro con la jerarquía provincia > cantón > distrito.

    La jerarquía no está codificada en columnas: las filas de cantón van
    en negrita; las de provincia se reconocen por su nombre, en el orden
    oficial; las demás son distritos. Devuelve una lista de tuplas
    (provincia, cantón, distrito o None, valor).
    """
    ws = libro[hoja]
    filas, provincia, canton, i_prov = [], None, None, 0
    for r in range(7, ws.max_row + 1):
        nombre, valor = ws.cell(r, 1).value, ws.cell(r, columna).value
        if nombre is None or not isinstance(valor, (int, float)) or nombre == "Costa Rica":
            continue
        nombre = nombre.strip()
        if ws.cell(r, 1).font.b:
            canton = nombre
            filas.append((provincia, canton, None, valor))
        elif i_prov < len(PROVINCIAS) and nombre == PROVINCIAS[i_prov]:
            provincia, i_prov = nombre, i_prov + 1
        else:
            filas.append((provincia, canton, nombre, valor))
    return filas


def main():
    with urllib.request.urlopen(URL) as respuesta:
        libro = openpyxl.load_workbook(io.BytesIO(respuesta.read()), data_only=True)
    poblacion = leer_cuadro(libro, "11", 2)
    area = leer_cuadro(libro, "12", 2)
    assert [p[:3] for p in poblacion] == [a[:3] for a in area], "los cuadros 11 y 12 no coinciden"

    # {(provincia, cantón): [población, área]} y datos de los distritos
    cantones, distritos = {}, {}
    for (prov, cant, dist, pob), (_, _, _, km2) in zip(poblacion, area):
        if dist is None:
            cantones[(prov, cant)] = [pob, km2]
        else:
            distritos[(prov, cant, dist)] = [pob, km2]
    assert len(cantones) == 82, f"se esperaban 82 cantones del INEC, hay {len(cantones)}"

    # Reclasificación de los distritos que hoy son cantones
    for prov, origen, nuevo in NUEVOS_CANTONES:
        pob, km2 = distritos[(prov, origen, nuevo)]
        cantones[(prov, origen)][0] -= pob
        cantones[(prov, origen)][1] -= km2
        cantones[(prov, nuevo)] = [pob, km2]
    assert len(cantones) == 84
    assert sum(v[0] for v in cantones.values()) == 5044197, "la población total no cuadra"

    # Orden del INEC (por provincia y cantón); los cantones nuevos, al final
    # de su provincia
    orden = sorted(cantones, key=lambda k: (PROVINCIAS.index(k[0]), list(cantones).index(k)))
    with open(SALIDA, "w", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow(["provincia", "canton", "poblacion", "area_km2"])
        for prov, cant in orden:
            pob, km2 = cantones[(prov, cant)]
            escritor.writerow([prov, cant, pob, f"{km2:.2f}"])
    print(f"{SALIDA}: {len(cantones)} cantones, "
          f"{sum(v[0] for v in cantones.values())} habitantes, "
          f"{sum(v[1] for v in cantones.values()):.2f} km²")


if __name__ == "__main__":
    main()
